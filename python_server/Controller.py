import mysql.connector
from openpyxl import load_workbook
from Product import Product
from Level import Level
from User import User
from Contract import Contract
from Client import Client
from ExecutedContract import ExecutedContract



class Controller:
    # connect to db server
    mydb = mysql.connector.connect(host="127.0.0.1", user="root", passwd="jst12345", database="myapp")

    cursor = mydb.cursor(buffered=True)

    mycursor = mydb.cursor()
    

    # ask for manager personal data and add manager (a.k.a. user without a mentor)
    # to the database
    def delete_user(self, user_id):
        sql = "DELETE FROM users WHERE user_id = %s"
        val = (user_id, )
        self.mycursor.execute(sql, val)
        self.mydb.commit()
    def delete_product(self, product_id):
        sql = "DELETE FROM products WHERE product_id = %s"
        val = (product_id, )
        self.mycursor.execute(sql, val)
        self.mydb.commit()
    def delete_client(self, client_id):
        sql = "DELETE FROM clients WHERE client_id = %s"
        val = (client_id, )
        self.mycursor.execute(sql, val)
        self.mydb.commit()
    def delete_level(self, level_id):
        sql = "DELETE FROM hierarchy WHERE level_id=%s"
        val = (level_id, )
        self.mycursor.excute(sql, val)
        self.mydb.commit()

    def add_manager(self, name_of_user, surname_of_user, personal_points, networking_points, paid_points, cash, paid_cash,
                    emso_of_user, municipality, geolocation_of_user, trr_of_user, bank_of_user, date, mentorship_date):

        sql = "SELECT name_of_user FROM users WHERE name_of_user=%s AND surname_of_user=%s"
        val = (name_of_user, surname_of_user)
        self.mycursor.execute(sql, val)
        duplicate = self.mycursor.fetchone()

        if duplicate is not None:
            pr_str = "ERROR: user with name %s and surname %s is already in the database. " % (
            name_of_user, surname_of_user)
            print(pr_str)
            return pr_str

        level = self.get_actual_level(personal_points+networking_points)

        if level is None:
            return "ERROR: Something is wrong with border points of levels. Could not find the level this manager should be in."


        sql = "INSERT INTO users(name_of_user, surname_of_user, personal_points, networking_points, paid_points, cash, paid_cash, level_id, EMSO, municipality, geolocation, TRR, bank, employment_contract_date, date_of_mentorship) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        val = (name_of_user, surname_of_user, personal_points, networking_points, paid_points, cash, paid_cash, level.level_id, emso_of_user,
               municipality, geolocation_of_user, trr_of_user, bank_of_user, date, mentorship_date)
        self.mycursor.execute(sql, val)
        self.mydb.commit()
        pr_str = "manager %s %s added" % (name_of_user, surname_of_user)
        print(pr_str)
        return "OK"


    # ask for user personal data and add user to the database
    def add_user(self, name_of_user, surname_of_user, mentor, personal_points, networking_points,
                     paid_points, cash, paid_cash,
                     emso_of_user, municipality, geolocation_of_user, trr_of_user,
                     bank_of_user, employment_contract_date_of_user, mentorship_date):

        print("Adding user")

        sql = "SELECT name_of_user FROM users WHERE name_of_user=%s AND surname_of_user=%s"
        val = (name_of_user, surname_of_user)
        self.mycursor.execute(sql, val)
        duplicate = self.mycursor.fetchone()

        warning = False

        level = self.get_actual_level(personal_points+networking_points)
        if level is None:
            pr_str = "ERROR: Something is wrong with the border points of the levels. Could not find the level this user should be in."
            print(pr_str)
            return pr_str

        if duplicate is not None:
            pr_str = "ERROR: user with name %s and surname %s is already in the database. " % (name_of_user, surname_of_user)
            print(pr_str)
            return pr_str

        if mentor is None:
            pr_str = "ERROR: The mentor %s %s of the user %s %s has not been added to the database yet. Add mentor with 'add_user' or 'add_manager' command and try again."
            print(pr_str)
            return pr_str
        if mentor.emso == emso_of_user and mentor.name == name_of_user and mentor.surname == surname_of_user:
            pr_str = "ERROR: Cannot add user that has himself as a mentor"
            print(pr_str)
            return pr_str

        if mentor.level.level_number <= level.level_number:
            pr_str = "WARNING: Mentor of user %s %s is not in higher level than him. Change the level of user or of the mentor " \
                     "so that mentor's level will be higher than user's level." % (name_of_user, surname_of_user)
            warning = True
            print(pr_str)


        sql = "INSERT INTO users(name_of_user, surname_of_user, mentor_id, personal_points, networking_points, paid_points, cash, paid_cash, level_id, EMSO, municipality, geolocation, TRR, bank, employment_contract_date, date_of_mentorship) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        val = (name_of_user, surname_of_user, mentor.id, personal_points, networking_points, paid_points, cash, paid_cash, level.level_id, emso_of_user,
               municipality, geolocation_of_user, trr_of_user, bank_of_user, employment_contract_date_of_user, mentorship_date)
        self.mycursor.execute(sql, val)
        self.mydb.commit()
        pr_str = "user %s %s added" % (name_of_user, surname_of_user)
        print(pr_str)
        if warning:
            return "WARNING: Mentor of user %s %s is not in higher level than him. Change the level of user or of the mentor " \
                     "so that mentor's level will be higher than user's level." % (name_of_user, surname_of_user)
        else:
            return "OK"


    # add a product to the database
    def add_product(self, name, partner, constant, plan_value, basic_name, plan_type, product_type, plan_name, bcv, points_earned, points_to_be_paid,
                    points_for_career, satbppr):

        sql = "SELECT product_name FROM products WHERE product_name=%s"
        val = (name,)
        self.mycursor.execute(sql, val)
        duplicate = self.mycursor.fetchone()

        if duplicate is not None:
            pr_str = "ERROR: Product with name %s is already in the database. " % (name)
            print(pr_str)
            return pr_str

        sql = "INSERT INTO products(product_name, product_partner, constant, plan_value, basic_name, plan_type, product_type, plan_name, bcv, points_earned, points_to_be_paid, points_for_career, satbppr) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
        val = (name, partner, constant, plan_value, basic_name, plan_type, product_type, plan_name, bcv, points_earned, points_to_be_paid,
                    points_for_career, satbppr)
        self.mycursor.execute(sql, val)
        self.mydb.commit()
        pr_str = "product %s added" % name
        print(pr_str)
        return "OK"


    # add a client to the database
    def add_client(self, name, surname, emso, address, city, email):

        sql = "SELECT client_name FROM clients WHERE client_name=%s AND client_surname=%s"
        val = (name, surname)
        self.mycursor.execute(sql, val)
        duplicate = self.mycursor.fetchone()

        if duplicate is not None:
            pr_str = "ERROR: Client with name %s and surname %s is already in the database. " % (name, surname)
            print(pr_str)
            return pr_str

        sql = "INSERT INTO clients(client_name, client_surname, EMSO, address, city, email) VALUES (%s, %s, %s, %s, %s, %s)"
        val = (name, surname, emso, address, city, email)
        self.mycursor.execute(sql, val)
        self.mydb.commit()
        pr_str = "client %s %s added" % (name, surname)
        print(pr_str)
        return "OK"


    # find user with given name and surname
    # return object of class user
    def find_user_from_name(self, name, surname):
        sql = "SELECT user_id, name_of_user, surname_of_user, mentor_id, personal_points, networking_points, paid_points, cash, paid_cash, level_id, emso, municipality, geolocation, TRR, bank, date_of_mentorship, employment_contract_date FROM users WHERE name_of_user = %s AND surname_of_user = %s"
        val = (name, surname)

        self.mycursor.execute(sql, val)

        user_data = self.mycursor.fetchone()

        if user_data is None:
            return None

        level = self.get_level_from_level_id(user_data[9])

        mentor = self.get_user_from_id(user_data[3])

        if mentor is None:
            return User(user_data[0], user_data[1], user_data[2], None, None,
                                None, user_data[4], user_data[5],
                                user_data[6], user_data[7], user_data[8],
                                level, user_data[10], user_data[11],
                                user_data[12], user_data[13], user_data[14], str(user_data[15]), str(user_data[16]))
        else:
            return User(user_data[0], user_data[1], user_data[2], mentor.name, mentor.surname,
                        user_data[3], user_data[4], user_data[5],
                        user_data[6], user_data[7], user_data[8],
                        level, user_data[10], user_data[11],
                        user_data[12], user_data[13], user_data[14], str(user_data[15]), str(user_data[16]))


    # add a contract and update the hierarchy (update points and cash of user who made
    # a contract and of his mentors)
    def add_contract(self, product_id, user_id,
                     client_id, date_of_payment,
                     base_value):

        user = self.get_user_from_id(user_id)

        if user is None:
            pr_str = "ERROR: The user %s %s has not been added to the database yet. Add the user with 'add_user' command and try again."
            print(pr_str)
            return pr_str

        client = self.get_client_from_id(client_id)

        if client is None:
            pr_str = "ERROR: The client %s %s has not been added to the database yet. Add the client with 'add_client' command and try again."
            print(pr_str)
            return pr_str

        product = self.get_product_from_id(product_id)

        if product is None:
            pr_str = "ERROR: The product %s has not been added to the database yet. Add the product with 'add_product' command and try again."
            print(pr_str)
            return pr_str

        contract = Contract(0, product, user, client, date_of_payment, base_value, float(base_value/product.constant),
                            self.get_level_from_level_id(user.level.level_id).in_eur, "NO")

        if not self.check_points_of_levels():
            pr_str = "ERROR: Something is wrong with border points of the levels. Contract not added to the database"
            print(pr_str)
            return pr_str

        # add a contract
        sql = "INSERT INTO contracts(product_id, user_id, client_id, date_of_payment, base_value, number_of_points, in_eur, paid_out) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"
        val = (contract.product.id, contract.user.id, contract.client.id, contract.date_of_payment, contract.base_value,
               contract.number_of_points, contract.in_eur, contract.paid_out)
        self.mycursor.execute(sql, val)
        self.mydb.commit()



        # update the hierarchy
        return self.update_hierarchy(product, user, client, contract)


        # pr_str = "contract made by user %s %s with product %s added" % (user.name, user.surname, product.product_name)
        # print(pr_str)
        # return "OK"


    # update hierarchy - update points and cash of user who sold the product and his mentors
    # and create executed contracts for each mentor
    # called after add_contract() function
    def update_hierarchy(self, product, user, client, contract):

        curr = user

        outgrownusers = []

        # update points and cash of all the superiors of the agent that made the contract and create executed contracts
        while True:

            # curr.level = get_level_number(curr)

            # calculate the points and cash the user "curr" will get
            if curr is user:
                points = float(contract.number_of_points)
                cash = points*curr.level.in_eur
                curr.personal_points = self.update_personal_points(curr, points, cash)
                curr.total_points = curr.personal_points + curr.networking_points
            else:
                points = contract.number_of_points
                cash = points * (curr.level.in_eur - inferior_level.in_eur)
                print("Name: "+curr.name+", level: "+str(curr.level.level_number)+", inferior_level: "+str(inferior_level.level_number))
                curr.networking_points = self.update_networking_points(curr, points, cash)
                curr.total_points = curr.personal_points + curr.networking_points

            # what level the curr_id should actually be in after updating the points
            actual_level = self.get_actual_level(curr.total_points)

            if actual_level is None:
                pr_str = "FATAL ERROR: Something is wrong with border points of the levels. Could not find the level user %s %s should be in. " % (curr.name, curr.surname)
                print(pr_str)
                return pr_str

            # get mentor of the current user
            mentor = self.get_user_from_id(curr.mentor_id)


            # check if curr needs to be promoted to another hierarchy level
            if curr.level.level_number != actual_level.level_number:

                self.set_level(curr, actual_level)

                # if superior of curr exists then check if curr needs a new superior
                if mentor is not None:

                    # check if curr needs a new mentor, eg if curr level is equal or more than his superior level
                    if mentor.level.level_number <= actual_level.level_number:
                        self.set_mentor(curr, self.get_manager())
                        outgrownusers.append(curr)


            # the loop is at the top of the pyramid -> break
            if mentor is None:
                break

            # add the executed contract for this superior
            val = (curr.mentor_id, curr.id, user.id, client.id, product.id, contract.date_of_payment, contract.base_value,
                   points, mentor.level.in_eur - curr.level.in_eur, "NO")
            self.mycursor.execute("INSERT INTO executed_contracts(user_id, subordinate_id, agent_id, client_id, product_id, date_of_payment, base_value, number_of_points, in_eur, paid_out) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)", val)

            inferior_level = curr.level
            curr = mentor

        self.mydb.commit()
        return outgrownusers


    def get_level_from_level_id(self, level_id):

        sql = "SELECT level_id, hierarchy_level, hierarchy_level_name, points_from, points_to, in_eur FROM hierarchy WHERE level_id = %s"
        val = (level_id, )
        self.mycursor.execute(sql, val)
        level_data = self.mycursor.fetchone()

        return Level(level_data[0], level_data[1], level_data[2], level_data[3], level_data[4], level_data[5])


    def get_manager(self):
        sql = "SELECT user_id, name_of_user, surname_of_user, mentor_id, personal_points, networking_points, paid_points, cash, paid_cash, level_id, emso, municipality, geolocation, TRR, bank, date_of_mentorship, employment_contract_date FROM users WHERE mentor_id IS NULL"
        self.mycursor.execute(sql)
        user_data = self.mycursor.fetchone()
        level = self.get_level_from_level_id(user_data[9])
        return User(user_data[0], user_data[1], user_data[2], None, None,
                    None, user_data[4], user_data[5],
                    user_data[6], user_data[7], user_data[8],
                    level, user_data[10], user_data[11],
                    user_data[12], user_data[13], user_data[14], str(user_data[15]), str(user_data[16]))

    # find a product with a given name
    # return object of class Product
    def find_product(self, name):

        sql = "SELECT product_id, product_name, product_partner, constant, plan_value, basic_name, plan_type, product_type, plan_name, bcv, points_earned, points_to_be_paid, points_for_career, satbppr FROM products WHERE product_name = %s"

        self.mycursor.execute(sql, (name, ))

        product_data = self.mycursor.fetchone()

        if product_data is None:
            return None

        return Product(product_data[0], product_data[1], product_data[2],
                       product_data[3], product_data[4], product_data[5],
                       product_data[6], product_data[7], product_data[8],
                       product_data[9], product_data[10], product_data[11],
                       product_data[12], product_data[13])


    # find client with a given name and surname
    # return object of class Product
    def find_client(self, name, surname):
        sql = "SELECT client_id, client_name, client_surname, EMSO, address, city, email FROM clients WHERE client_name = %s AND client_surname = %s"
        val = (name, surname)

        self.mycursor.execute(sql, val)
        client_data = self.mycursor.fetchone()

        if client_data is None:
            return None

        return Client(client_data[0], client_data[1], client_data[2], client_data[3], client_data[4], client_data[5], client_data[6])


    # set the level of given user
    def set_level(self, user, level):

        # update hierarchy level of the user
        sql = "UPDATE users SET level_id = %s WHERE user_id = %s"
        self.mycursor.execute(sql, (level.level_id, user.id))
        self.mydb.commit()

        return "OK"


    # calculate what level should an user actually be in depending on the points he has
    def get_actual_level(self, points):

        # check what level should he actually be in
        sql = "SELECT level_id, hierarchy_level, hierarchy_level_name, points_from, points_to, in_eur FROM hierarchy WHERE points_from <= %s AND points_to > %s OR hierarchy_level = (SELECT MAX(hierarchy_level) FROM hierarchy) AND points_from < %s"
        self.mycursor.execute(sql, (points, points, points))
        level_data = self.mycursor.fetchone()
        return Level(level_data[0], level_data[1], level_data[2], level_data[3], level_data[4], level_data[5])


    # update points AND cash of the user
    def update_personal_points(self, user, points, cash):

        # get current points and cash
        sql = "SELECT personal_points, cash FROM users WHERE user_id = %s"
        self.mycursor.execute(sql, (user.id, ))

        pts_cash = self.mycursor.fetchone()
        pts = float(pts_cash[0])
        csh = float(pts_cash[1])

        sql = "UPDATE users SET personal_points = %s, cash  =%s WHERE user_id = %s"
        self.mycursor.execute(sql, (pts + points, cash+csh, user.id))
        self.mydb.commit()

        return pts + points


    def update_networking_points(self, user, points, cash):
        # get current points and cash
        sql = "SELECT networking_points, cash FROM users WHERE user_id = %s"
        self.mycursor.execute(sql, (user.id,))

        pts_cash = self.mycursor.fetchone()
        pts = float(pts_cash[0])
        csh = float(pts_cash[1])

        sql = "UPDATE users SET networking_points = %s, cash=%s WHERE user_id = %s"
        self.mycursor.execute(sql, (pts + points, csh+cash, user.id))
        self.mydb.commit()

        return pts + points


    # set the mentor of the user
    def set_mentor(self, user, mentor):

        if user.id == mentor.id:
            pr_str = "User cannot be mentor of himself"
            print(pr_str)
            return pr_str

        sql = "UPDATE users SET mentor_id = %s WHERE user_id = %s"
        self.mycursor.execute(sql, (mentor.id, user.id))
        self.mydb.commit()
        return "OK"


    # ask the manager for the new mentor (called when the user needs a new mentor)
    def ask_for_new_mentor(self, user, level):

        string = "user %s %s with ID: %s has been promoted to level %s and needs a new mentor." % (user.name, user.surname, user.id, level.level_number)
        print(string)

        while True:
            mentor_name = str(input("Enter name of the new mentor: "))
            mentor_surname = str(input("Enter surname of the new mentor: "))

            mentor = self.find_user_from_name(mentor_name, mentor_surname)
            if mentor is None:
                string = "user %s %s has not been added to the database yet and therefore cannot be mentor. Add him with 'add_manager' or 'add_user' command or try again." % (mentor_name, mentor_surname)
                print(string)
            else:
                break

        return mentor


    # change personal data of the user (currently name, surname and trr are supported)
    def change_user(self, user_id, what_to_change, new_value1):

        pr_str = "OK"

        # update the name of the user
        if what_to_change == "name":
            sql = "UPDATE users SET name_of_user = %s WHERE user_id = %s"
            self.mycursor.execute(sql, (new_value1, user_id))

        elif what_to_change == "municipality":
            sql = "UPDATE users SET municipality = %s WHERE user_id = %s"
            self.mycursor.execute(sql, (new_value1, user_id))

        elif what_to_change == "surname":
            sql = "UPDATE users SET surname_of_user = %s WHERE user_id = %s"
            self.mycursor.execute(sql, (new_value1, user_id))

        elif what_to_change == "mentor":
            if new_value1 == user_id:
                pr_str = "User cannot be mentor of himself"
            else:
                sql = "UPDATE users SET mentor_id = %s WHERE user_id = %s"
                self.mycursor.execute(sql, (new_value1, user_id))

        elif what_to_change == "date_of_mentorship":
            sql = "UPDATE users SET date_of_mentorship = %s WHERE user_id = %s"
            self.mycursor.execute(sql, (new_value1, user_id))

        elif what_to_change == "points":
            pr_str = "Not supported"
        elif what_to_change == "cash":
            pr_str = "Not supported"
        elif what_to_change == "level_number":
            pr_str = "Not supported"

        elif what_to_change == "emso":
            sql = "UPDATE users SET emso = %s WHERE user_id = %s"
            self.mycursor.execute(sql, (new_value1, user_id))

        elif what_to_change == "geolocation":
            sql = "UPDATE users SET geolocation = %s WHERE user_id = %s"
            self.mycursor.execute(sql, (new_value1, user_id))

        elif what_to_change == "TRR":
            sql = "UPDATE users SET TRR = %s WHERE user_id = %s"
            self.mycursor.execute(sql, (new_value1, user_id))

        elif what_to_change == "bank":
            sql = "UPDATE users SET bank = %s WHERE user_id = %s"
            self.mycursor.execute(sql, (new_value1, user_id))

        elif what_to_change == "employment_contract_date":
            print(new_value1)
            sql = "UPDATE users SET employment_contract_date = %s WHERE user_id = %s"
            self.mycursor.execute(sql, (new_value1, user_id))

        else:
            pr_str = "Cannot change this attribute"

        self.mydb.commit()
        print(pr_str)
        return pr_str


    def change_product(self, product_id, what_to_change, new_value):

        pr_str = "OK"

        if what_to_change == "name":
            sql = "UPDATE products SET product_name = %s WHERE product_id = %s"
            self.mycursor.execute(sql, (new_value, product_id))

        elif what_to_change == "product_partner":
            sql = "UPDATE products SET product_partner = %s WHERE product_id = %s"
            self.mycursor.execute(sql, (new_value, product_id))

        elif what_to_change == "constant":
            sql = "UPDATE products SET constant = %s WHERE product_id = %s"
            self.mycursor.execute(sql, (new_value, product_id))

        elif what_to_change == "plan_value":
            sql = "UPDATE products SET plan_value = %s WHERE product_id = %s"
            self.mycursor.execute(sql, (new_value, product_id))

        elif what_to_change == "basic_name":
            sql = "UPDATE products SET basic_name = %s WHERE product_id = %s"
            self.mycursor.execute(sql, (new_value, product_id))

        elif what_to_change == "plan_type":
            sql = "UPDATE products SET plan_type = %s WHERE product_id = %s"
            self.mycursor.execute(sql, (new_value, product_id))

        else:
            pr_str = "Cannot change this attribute"

        self.mydb.commit()
        print(pr_str)
        return pr_str


    # change data of the contract - not supported
    def change_contract(self):

        pr_str = "Not supported"
        return pr_str


    def change_executed_contract(self):

        pr_str = "Not supported"
        return pr_str


    def get_level_from_level_number(self, level_number):

        sql = "SELECT level_id, hierarchy_level, hierarchy_level_name, points_from, points_to, in_eur FROM hierarchy WHERE hierarchy_level = %s"
        val = (level_number, )

        self.mycursor.execute(sql, val)
        level_data = self.mycursor.fetchone()

        if level_data is None:
            return None

        return Level(level_data[0], level_data[1], level_data[2], level_data[3], level_data[4], level_data[5])


    def change_level(self, level_id, what_to_change, new_value):

        pr_str = "OK"

        level = self.get_level_from_level_id(level_id)
        if level is None:
            pr_str = "No such level"
            print(pr_str)
            return pr_str

        if what_to_change == "level_number":
            sql = "UPDATE hierarchy SET hierarchy_level = %s WHERE level_id = %s"
            self.mycursor.execute(sql, (new_value, level.level_id))

        elif what_to_change == "level_name":
            sql = "UPDATE hierarchy SET hierarchy_level_name = %s WHERE level_id = %s"
            self.mycursor.execute(sql, (new_value, level.level_id))

        elif what_to_change == "points_from":
            sql = "UPDATE hierarchy SET points_from = %s WHERE level_id = %s"
            self.mycursor.execute(sql, (new_value, level.level_id))

            if self.check_points_of_levels():
                self.update_levels_of_users()

        elif what_to_change == "points_to":
            sql = "UPDATE hierarchy SET points_to = %s WHERE level_id = %s"
            self.mycursor.execute(sql, (new_value, level.level_id))

            if self.check_points_of_levels():
                self.update_levels_of_users()


        elif what_to_change == "in_eur":
            sql = "UPDATE hierarchy SET in_eur = %s WHERE level_id = %s"
            self.mycursor.execute(sql, (new_value, level.level_id))

        else:
            pr_str = "Cannot update this attribute"
            print(pr_str)
            return pr_str

        self.mydb.commit()
        return pr_str

    # go through all users and check if they need to be in different levels
    def update_levels_of_users(self):

        sql = "SELECT user_id, name_of_user, surname_of_user, mentor_id, personal_points, networking_points, paid_points, cash, paid_cash, level_id, emso, municipality, geolocation, TRR, bank, date_of_mentorship, employment_contract_date FROM users"
        self.mycursor.execute(sql)
        users = self.mycursor.fetchall()

        print("Updating levels of users ...")

        for user_data in users:


            level = self.get_level_from_level_id(user_data[9])
            mentor = self.get_user_from_id(user_data[3])

            if mentor is None:
                curr = User(user_data[0], user_data[1], user_data[2], None, None, None, user_data[4],
                            user_data[5], user_data[6], user_data[7], user_data[8], level,
                            user_data[10], user_data[11], user_data[12], user_data[13], user_data[14], str(user_data[15]), str(user_data[16]))
            else:
                curr = User(user_data[0], user_data[1], user_data[2], mentor.name, mentor.surname, mentor.id, user_data[4],
                            user_data[5], user_data[6], user_data[7], user_data[8], level,
                            user_data[10], user_data[11], user_data[12], user_data[13], user_data[14], str(user_data[15]), str(user_data[16]))


            actual_level = self.get_actual_level(curr.total_points)

            if curr.level.level_id != actual_level.level_id:
                self.set_level(curr, actual_level)
                pr_str = "    %s %s updated to level %s" % (curr.name, curr.surname, actual_level.level_number)
                print(pr_str)

        print("    Done")
        return "OK"


    # print contracts from and to input date
    def show_contracts(self, date_from, date_to):

        sql = "SELECT contract_id, user_id, product_id, date_of_payment, number_of_points, paid_out, client_id, base_value, in_eur FROM contracts WHERE date_of_payment >= %s AND date_of_payment <= %s"
        self.mycursor.execute(sql, (date_from, date_to))
        contracts = self.mycursor.fetchall()

        i = 0

        for contract_data in contracts:
            user_id = contract_data[1]
            product_id = contract_data[2]
            client_id = contract_data[6]

            user = self.get_user_from_id(user_id)
            product = self.get_product_from_id(product_id)
            client = self.get_client_from_id(client_id)

            pr_str = "%s %s %s, %s, %s | PERSONAL POINTS: %s, PAID: %s" % (contract_data[0], user.name, user.surname, product.product_name, contract_data[3], contract_data[4], contract_data[5])
            print(pr_str)

            # Contract(contract_id, product, user, client, date_of_payment, base_value, number_of_points, in_eur, paid_out)
            contracts[i] = Contract(contract_data[0], product, user, client, str(contract_data[3]), contract_data[7], contract_data[4], contract_data[8], contract_data[5])

            i = i + 1

        return contracts


    # print executed contracts
    def show_executed_contracts(self, date_from, date_to):

        sql = "SELECT contract_id, user_id, product_id, date_of_payment, number_of_points, paid_out, subordinate_id, client_id, base_value, in_eur FROM executed_contracts WHERE date_of_payment >= %s AND date_of_payment <= %s"
        self.mycursor.execute(sql, (date_from, date_to))
        executed_contracts = self.mycursor.fetchall()

        i = 0

        for executed_contract_data in executed_contracts:
            user_id = executed_contract_data[1]
            product_id = executed_contract_data[2]

            user = self.get_user_from_id(user_id)
            subordinate = self.get_user_from_id(executed_contract_data[6])
            client = self.get_client_from_id(executed_contract_data[7])

            product = self.get_product_from_id(product_id)

            pr_str = "%s %s %s, %s, %s | NETWORKING POINTS: %s, PAID: %s" % (executed_contract_data[0], user.name, user.surname, product.product_name, executed_contract_data[3], executed_contract_data[4], executed_contract_data[5])
            print(pr_str)

            # ExecutedContract(contract_id, product, user, subordinate, client, date_of_payment, base_value, number_of_points, in_eur, paid_out)
            executed_contracts[i] = ExecutedContract(executed_contract_data[0], product,
                                                    user, subordinate, client, str(executed_contract_data[3]),
                                                    executed_contract_data[8], executed_contract_data[4],
                                                    executed_contract_data[9], executed_contract_data[5])

            i = i + 1

        return executed_contracts


    def get_user_from_id(self, user_id):

        if user_id is None:
            return None

        sql = "SELECT user_id, name_of_user, surname_of_user, mentor_id, personal_points, networking_points, paid_points, cash, paid_cash, level_id, emso, municipality, geolocation, TRR, bank, date_of_mentorship, employment_contract_date FROM users WHERE user_id = %s"
        val = (user_id, )
        self.mycursor.execute(sql, val)

        user_data = self.mycursor.fetchone()

        if user_data is None:
            return None

        level = self.get_level_from_level_id(user_data[9])

        if user_data[3] is None:
            return User(user_data[0], user_data[1], user_data[2], None, None,
                            None, user_data[4], user_data[5],
                            user_data[6], user_data[7], user_data[8],
                            level, user_data[10], user_data[11],
                            user_data[12], user_data[13], user_data[14], str(user_data[15]), str(user_data[16]))
        else:
            # get mentor name and surname
            sql = "SELECT name_of_user, surname_of_user FROM users WHERE user_id = %s"
            val = (user_data[3],)
            self.mycursor.execute(sql, val)
            mentor_data = self.mycursor.fetchone()

            return User(user_data[0], user_data[1], user_data[2], mentor_data[0], mentor_data[1],
                        user_data[3], user_data[4], user_data[5],
                        user_data[6], user_data[7], user_data[8],
                        level, user_data[10], user_data[11],
                        user_data[12], user_data[13], user_data[14], str(user_data[15]), str(user_data[16]))


    def get_product_from_id(self, product_id):
        sql = "SELECT product_id, product_name, product_partner, constant, plan_value, basic_name, plan_type, product_type, plan_name, bcv, points_earned, points_to_be_paid, points_for_career, satbppr FROM products WHERE product_id = %s"
        val = (product_id, )

        self.mycursor.execute(sql, val)

        product_data = self.mycursor.fetchone()

        if product_data is None:
            return None

        return Product(product_data[0], product_data[1], product_data[2],
                       product_data[3], product_data[4], product_data[5],
                       product_data[6], product_data[7], product_data[8],
                       product_data[9], product_data[10], product_data[11],
                       product_data[12], product_data[13])


    # add a hierarchy level - must be called first if the database is empty
    def add_level(self, level_number, level_name, points_from, points_to, in_eur):

        # find possible duplicate
        sql = "SELECT level_id FROM hierarchy WHERE hierarchy_level = %s OR hierarchy_level_name = %s"
        val = (level_number, level_name)
        self.mycursor.execute(sql, val)
        possible_duplicate = self.mycursor.fetchone()

        # duplicate exists
        if possible_duplicate is not None:
            # string = "Level %s, %s already exists, do you want to override this level with new data" % (level_number, level_name)
            # print(string)
            # answer = str(input("Y/N:"))
            #
            # if answer == "y" or answer == "Y":
            #     sql = "UPDATE hierarchy SET hierarchy_level = %s, hierarchy_level_name = %s, points_from = %s, points_to = %s, in_eur = %s WHERE hierarchy_level = %s OR hierarchy_level_name = %s"
            #     val = (level_number, name, points_from, points_to, in_eur, level_number, level_name)
            #     self.mycursor.execute(sql, val)
            #     self.mydb.commit()
            # else:
            #     print("Did not update")
            pr_str = "ERROR: Level %s, %s already exists, try changing this level, or add new level with different data" % (level_number, level_name)
            print(pr_str)

            return pr_str

        sql = "INSERT INTO hierarchy(hierarchy_level, hierarchy_level_name, points_from, points_to, in_eur) VALUES (%s, %s, %s, %s, %s)"
        val = (level_number, level_name, points_from, points_to, in_eur)
        self.mycursor.execute(sql, val)
        self.mydb.commit()
        pr_str = "hierarchy level %s added" % (level_number)
        print(pr_str)

        if self.check_points_of_levels():
            self.update_levels_of_users()

        return "OK"


    # check if level's points_from and points_to are set correctly
    def check_points_of_levels(self):

        print("Checking border points of all levels ...")

        sql = "SELECT points_from, points_to, hierarchy_level FROM hierarchy"
        self.mycursor.execute(sql)
        checking_levels = self.mycursor.fetchall()

        prev_points_to = 0
        sth_wrong = False

        for checking_level in checking_levels:

            if checking_level[0] != prev_points_to:
                pr_str = "    WARNING: Border points of level %s do not match its neighbouring levels. Check POINTS FROM and POINTS TO of this level and change or add level with 'change_level' or 'add_level' command." % checking_level[2]
                print(pr_str)
                sth_wrong = True

            prev_points_to = checking_level[1]

        if sth_wrong:
            return False
        else:
            print("    All OK")
            return True


    # show hierarchy levels
    def show_levels(self):

        sql = "SELECT hierarchy_level, hierarchy_level_name, points_from, points_to, in_eur, level_id FROM hierarchy"
        self.mycursor.execute(sql)
        levels = self.mycursor.fetchall()

        i = 0

        for level_data in levels:
            pr_str = "%s %s: %s-%s | EUR VALUE: %s" % (level_data[0], level_data[1], level_data[2], level_data[3], level_data[4])
            print(pr_str)

            # Level(level_id, hierarchy_level, hierarchy_name, points_from, points_to, in_eur)
            levels[i] = Level(level_data[5], level_data[0], level_data[1], level_data[2], level_data[3], level_data[4])

            i = i + 1

        return levels


    def get_contract_from_contract_id(self, contract_id):
        sql = "SELECT contract_id, product_id, user_id, client_id, date_of_payment, base_value, number_of_points, in_eur, paid_out FROM contracts WHERE contract_id = %s"
        val = (contract_id, )
        self.mycursor.execute(sql, val)
        contract = self.mycursor.fetchone()

        if contract is None:
            return None

        product = self.get_product_from_id(contract[1])
        user = self.get_user_from_id(contract[2])
        client = self.get_client_from_id(contract[3])

        return Contract(contract[0], product, user,
                                             client, str(contract[4]),
                                             contract[5], contract[6], contract[7],
                                             contract[8])


    def get_executed_contract_from_contract_id(self, contract_id):
        sql = "SELECT contract_id, product_id, user_id, subordinate_id, client_id, date_of_payment, base_value, number_of_points, in_eur, paid_out FROM executed_contracts WHERE contract_id = %s"
        val = (contract_id,)
        self.mycursor.execute(sql, val)
        executed_contract = self.mycursor.fetchone()

        if executed_contract is None:
            return None

        product = self.get_product_from_id(executed_contract[1])
        user = self.get_user_from_id(executed_contract[2])
        subordinate = self.get_user_from_id(executed_contract[3])
        client = self.get_client_from_id(executed_contract[4])

        return ExecutedContract(executed_contract[0], product, user, subordinate, client,
                                             str(executed_contract[5]), executed_contract[6], executed_contract[7], executed_contract[8], executed_contract[9])





    # show products
    def show_products(self):

        sql = "SELECT product_id, product_name, product_partner, constant, plan_value, basic_name, plan_type, product_type, plan_name, bcv, points_earned, points_to_be_paid, points_for_career, satbppr FROM products"
        self.mycursor.execute(sql)
        products = self.mycursor.fetchall()

        i = 0

        for product_data in products:

            # Product(id, product_name, product_partner,
            #                  constant, plan_value, basic_name, plan_type,
            #                  product_type, plan_name, bcv, points_earned,
            #                  points_to_be_paid, points_for_carrer,
            #                  satbppr)
            print(product_data[1], product_data[2], product_data[3])
            products[i] = Product(product_data[0], product_data[1], product_data[2], product_data[3], product_data[4], product_data[5], product_data[6],
                                  product_data[7], product_data[8], product_data[9], product_data[10], product_data[11], product_data[12], product_data[13])

            i = i + 1

        return products


    # show users
    def show_users(self):
        sql = "SELECT name_of_user, surname_of_user, level_id, personal_points, networking_points, paid_points, cash, paid_cash, mentor_id, user_id, emso, municipality, geolocation, trr, bank, date_of_mentorship, employment_contract_date FROM users"
        self.mycursor.execute(sql)
        users = self.mycursor.fetchall()

        i = 0

        print(users)

        # User(id, name, surname, mentor, personal_points, networking_points, paid_points,
        #                 cash, paid_cash, level, emso, municipality, geolocation, trr, bank)



        for user_data in users:
            level = self.get_level_from_level_id(user_data[2])
            mentor = self.get_user_from_id(user_data[8])

            if mentor is None:
                string = "%s %s, level: %s, personal points: %s, networking points: %s, total_points: %s, paid points: %s, cash: %s, paid_cash: %s, No Mentor" % (
                    user_data[0], user_data[1], level.level_number, user_data[3], user_data[4], user_data[3]+user_data[4], user_data[5], user_data[6], user_data[7])
                users[i] = User(user_data[9], user_data[0], user_data[1], None, None, None,
                                user_data[3], user_data[4],
                                user_data[5], user_data[6], user_data[7], level, user_data[10], user_data[11],
                                user_data[12],
                                user_data[13], user_data[14], str(user_data[15]), str(user_data[16]))
                print(string)
            else:
                string = "%s %s, level: %s, personal points: %s, networking points: %s, total_points: %s, paid points: %s, cash: %s, paid_cash: %s, mentor: %s %s" % (
                    user_data[0], user_data[1], level.level_number, user_data[3], user_data[4], user_data[3]+user_data[4], user_data[5], user_data[6], user_data[7],
                    mentor.name, mentor.surname)
                users[i] = User(user_data[9], user_data[0], user_data[1], mentor.name, mentor.surname, mentor.id,
                                user_data[3], user_data[4],
                                user_data[5], user_data[6], user_data[7], level, user_data[10], user_data[11],
                                user_data[12],
                                user_data[13], user_data[14], str(user_data[15]), str(user_data[16]))
                print(string)



            i = i + 1

        return users

    # show clients
    def show_clients(self):

        sql = "SELECT client_id, client_name, client_surname, emso, address, city, email FROM clients"
        self.mycursor.execute(sql)
        clients = self.mycursor.fetchall()

        i = 0

        for client_data in clients:

            print(client_data[1]+" "+client_data[2])

            # Client(id, name, surname, emso, address, city, email)

            clients[i] = Client(client_data[0], client_data[1], client_data[2], client_data[3],
                                client_data[4], client_data[5], client_data[6])

            i = i + 1

        return clients


    def add_levels_from_table(self):
        levels = load_workbook(filename="levels.xlsx")
        levels_sheet = levels.active

        row = 2

        while True:
            level_number = levels_sheet.cell(row=row, column=2).value

            if level_number is None:
                break

            level_number = int(level_number)
            level_name = str(levels_sheet.cell(row=row, column=3).value)
            points_from = float(levels_sheet.cell(row=row, column=4).value)
            points_to = float(levels_sheet.cell(row=row, column=5).value)
            in_eur = float(levels_sheet.cell(row=row, column=6).value)

            self.add_level(level_number, level_name, points_from, points_to, in_eur)
            row = row + 1


    def add_users_from_table(self):
        users = load_workbook(filename="users.xlsx")
        users_sheet = users.active

        row = 2

        while True:
            user_name = users_sheet.cell(row=row, column=2).value

            if user_name is None:
                break

            user_name = str(user_name)
            user_surname = str(users_sheet.cell(row=row, column=3).value)
            mentor_name = users_sheet.cell(row=row, column=4).value
            mentor_surname = users_sheet.cell(row=row, column=5).value
            personal_points = float(users_sheet.cell(row=row, column=6).value)
            networking_points = float(users_sheet.cell(row=row, column=7).value)
            cash = float(users_sheet.cell(row=row, column=9).value)
            level_number = int(users_sheet.cell(row=row, column=10).value)
            emso = int(users_sheet.cell(row=row, column=11).value)
            municipality = str(users_sheet.cell(row=row, column=12).value)
            geolocation = str(users_sheet.cell(row=row, column=13).value)
            trr = str(users_sheet.cell(row=row, column=14).value)
            bank = str(users_sheet.cell(row=row, column=15).value)
            date = str(users_sheet.cell(row=row, column=16).value)

            if mentor_name is None:
                self.add_manager(user_name, user_surname, personal_points, networking_points, personal_points+networking_points, cash, cash, emso,
                            municipality, geolocation, trr, bank, date, 0)
            else:
                mentor_name = str(mentor_name)
                mentor_surname = str(mentor_surname)

                mentor = self.find_user_from_name(mentor_name, mentor_surname)

                self.add_user(user_name, user_surname, mentor, personal_points, networking_points, personal_points+networking_points, cash, cash, emso,
                             municipality, geolocation, trr, bank, date, 0)

            row = row + 1


    def add_clients_from_table(self):
        clients = load_workbook(filename="clients.xlsx")
        clients_sheet = clients.active

        row = 2

        while True:
            name = clients_sheet.cell(row=row, column=2).value

            if name is None:
                break


            name = str(name)
            surname = str(clients_sheet.cell(row=row, column=3).value)
            emso = int(clients_sheet.cell(row=row, column=4).value)
            address = str(clients_sheet.cell(row=row, column=5).value)
            city = str(clients_sheet.cell(row=row, column=6).value)
            email = str(clients_sheet.cell(row=row, column=7).value)

            self.add_client(name, surname, emso, address, city, email)
            row = row + 1


    def add_products_from_table(self):
        products = load_workbook(filename="products.xlsx")
        products_sheet = products.active

        row = 2

        while True:

            name = products_sheet.cell(row = row, column=4).value

            if name is None:
                break

            name = str(name)
            product_type = str(products_sheet.cell(row = row, column=2).value)
            partner = str(products_sheet.cell(row = row, column=3).value)
            plan_name = str(products_sheet.cell(row = row, column=5).value)
            constant = float(products_sheet.cell(row = row, column=6).value)
            bcv = float(products_sheet.cell(row = row, column=7).value)
            points_to_be_paid = str(products_sheet.cell(row = row, column=9).value)
            points_for_career = str(products_sheet.cell(row = row, column=10).value)
            satbppr = str(products_sheet.cell(row = row, column=11).value)

            self.add_product(name, partner, constant, 0, "", "",
                        product_type, plan_name, bcv, bcv/constant, points_to_be_paid,
                        points_for_career, satbppr)

            row = row + 1


    def change_client(self, client_id, what_to_change, new_value):

        if what_to_change == "name":
            sql = "UPDATE clients SET client_name = %s WHERE client_id = %s"
            val = (new_value, client_id)
            self.mycursor.execute(sql, val)

        elif what_to_change == "surname":
            sql = "UPDATE clients SET client_surname = %s WHERE client_id = %s"
            val = (new_value, client_id)
            self.mycursor.execute(sql, val)

        elif what_to_change == "emso":
            sql = "UPDATE clients SET EMSO = %s WHERE client_id = %s"
            val = (new_value, client_id)
            self.mycursor.execute(sql, val)

        self.mydb.commit()

        return "OK"


    # add data from table
    # you can comment the function of the data you dont want to add from table here
    def add_from_table(self):

        self.add_levels_from_table()
        self.add_users_from_table()
        self.add_clients_from_table()
        self.add_products_from_table()


    #pay out cash from points
    def pay_out_contracts(self, user, date_from, date_to):

        if user is None:
            pr_str = "ERROR: This user is not in the database yet."
            print(pr_str)
            return pr_str

        pr_str = "These are contracts of user %s %s from %s to %s that have not been paid out yet: \n" % (user.name, user.surname, date_from, date_to)
        print(pr_str)

        contracts = self.show_contracts_of(user, date_from, date_to)

        if len(contracts) == 0:
            pr_str = "There are no contracts of this user."
            print(pr_str)
            return pr_str

        indeces = input("Enter indeces of contracts you would like to pay out (separated by ' '): ").split(" ")

        paid_points_to_add = 0
        cash_to_add = 0

        # chosen contracts
        for i in indeces:

            paid_points_to_add = paid_points_to_add + float(contracts[int(i)].number_of_points)
            cash_to_add = cash_to_add + (float(contracts[int(i)].number_of_points) * float(contracts[int(i)].in_eur))

            sql = "UPDATE contracts SET paid_out=%s WHERE contract_id = %s"
            val = ("YES", contracts[int(i)].contract_id)
            self.mycursor.execute(sql, val)

        sql = "UPDATE users SET paid_points=%s, paid_cash=%s WHERE user_id = %s"
        val = (user.paid_points + paid_points_to_add, user.paid_cash + cash_to_add, user.id)
        self.mycursor.execute(sql, val)
        user.paid_points = user.paid_points + paid_points_to_add
        user.cash = user.paid_cash + cash_to_add
        self.mydb.commit()

        return "OK"


    def show_executed_contracts_of(self, user, date_from, date_to):
        sql = "SELECT contract_id, product_id, user_id, subordinate_id, client_id, date_of_payment, base_value, number_of_points, in_eur, paid_out FROM executed_contracts WHERE date_of_payment>=%s AND date_of_payment<=%s AND user_id = %s AND paid_out=%s"
        val = (date_from, date_to, user.id, "NO")
        self.mycursor.execute(sql, val)
        executed_contracts = self.mycursor.fetchall()

        i = 0

        for executed_contract in executed_contracts:

            product = self.get_product_from_id(executed_contracts[i].product_id)
            user = self.get_user_from_id(executed_contracts[i].user_id)
            subordinate = self.get_user_from_id(executed_contracts[i].subordinate_id)
            client = self.get_client_from_id(executed_contracts[i].client_id)

            executed_contracts[i] = ExecutedContract(executed_contract[0], product, user, subordinate, client,
                                             executed_contract[5], executed_contract[6], executed_contract[7], executed_contract[8], executed_contract[9])


            pr_str = "Index: %s, Product: %s, user: %s %s, Date: %s, Points: %s | PAID: %s" % (
                                                                    i, product.product_name, user.name,
                                                                    user.surname, executed_contracts[i].date_of_payment,
                                                                    executed_contracts[i].number_of_points, executed_contracts[i].paid_out)

            print(pr_str)

            i = i + 1

        return executed_contracts


    def get_client_from_id(self, client_id):
        sql = "SELECT client_id, client_name, client_surname, address, city, email, emso FROM clients WHERE client_id = %s"
        val = (client_id, )
        self.mycursor.execute(sql, val)
        client_data = self.mycursor.fetchone()

        return Client(client_data[0], client_data[1], client_data[2],
                      client_data[3], client_data[4], client_data[5],
                      client_data[6])


    def show_contracts_of(self, user, date_from, date_to):
        sql = "SELECT contract_id, product_id, user_id, client_id, date_of_payment, base_value, number_of_points, in_eur, paid_out FROM contracts WHERE date_of_payment>=%s AND date_of_payment<=%s AND user_id = %s AND paid_out=%s"
        val = (date_from, date_to, user.id, "NO")
        self.mycursor.execute(sql, val)
        contracts = self.mycursor.fetchall()

        i = 0
        for contract in contracts:

            product = self.get_product_from_id(contracts[i].product_id)
            user = self.get_user_from_id(contracts[i].user_id)
            client = self.get_client_from_id(contracts[i].client_id)

            contracts[i] = Contract(contract[0], product, user,
                                             client, contract[4],
                                             contract[5], contract[6], contract[7],
                                             contract[8])


            pr_str = "Index: %s, Product: %s, user: %s %s, Date: %s, Points: %s | PAID: %s" % (i, product.product_name, user.name,
                                                                                                   user.surname, contracts[i].date_of_payment,
                                                                                                   contracts[i].number_of_points, contracts[i].paid_out)
            print(pr_str)

            i = i + 1

        return contracts

    def pay_out_executed_contract(self, contract):

        print("paying out executed contract")

        check_ex_contract = self.get_executed_contract_from_contract_id(contract.contract_id)
        if check_ex_contract.paid_out == "YES":
            return


        paid_points_to_add = float(contract.number_of_points)
        cash_to_add = float(contract.number_of_points) * float(contract.in_eur)

        sql = "UPDATE users SET paid_points=%s, paid_cash=%s WHERE user_id = %s"
        val = (contract.user.paid_points + paid_points_to_add, contract.user.paid_cash + cash_to_add, contract.user.id)
        self.mycursor.execute(sql, val)
        contract.user.paid_points = contract.user.paid_points + paid_points_to_add
        contract.user.cash = contract.user.paid_cash + cash_to_add

        sql = "UPDATE executed_contracts SET paid_out=%s WHERE contract_id = %s"
        val = ("YES", contract.contract_id)
        self.mycursor.execute(sql, val)

        self.mydb.commit()

    def pay_out_contract(self, contract):

        check_contract = self.get_contract_from_contract_id(contract.contract_id)
        if check_contract.paid_out == "YES":
            return

        paid_points_to_add = float(contract.number_of_points)
        cash_to_add = float(contract.number_of_points) * float(contract.in_eur)

        sql = "UPDATE users SET paid_points=%s, paid_cash=%s WHERE user_id = %s"
        val = (contract.user.paid_points + paid_points_to_add, contract.user.paid_cash + cash_to_add, contract.user.id)
        self.mycursor.execute(sql, val)
        contract.user.paid_points = contract.user.paid_points + paid_points_to_add
        contract.user.cash = contract.user.paid_cash + cash_to_add

        sql = "UPDATE contracts SET paid_out=%s WHERE contract_id = %s"
        val = ("YES", contract.contract_id)
        self.mycursor.execute(sql, val)

        self.mydb.commit()


    def pay_out_executed_contracts(self, user, date_from, date_to):

        if user is None:
            pr_str = "ERROR: This user is not in the database yet."
            print(pr_str)
            return pr_str

        pr_str = "These are executed contracts of user %s %s from %s to %s that have not been paid out yet: " % (
            user.name, user.surname, date_from, date_to)
        print(pr_str)

        executed_contracts = self.show_executed_contracts_of(user, date_from, date_to)

        if len(executed_contracts) == 0:
            print("There are no executed contracts of this user.")
            return

        indeces = input("Enter indeces of executed contracts you would like to pay out (separated by ' '): ").split(" ")

        paid_points_to_add = 0
        cash_to_add = 0

        for i in indeces:
            paid_points_to_add = paid_points_to_add + float(executed_contracts[int(i)].number_of_points)
            cash_to_add = cash_to_add + (
                        float(executed_contracts[int(i)].number_of_points) * float(executed_contracts[int(i)].in_eur))

            sql = "UPDATE executed_contracts SET paid_out=%s WHERE contract_id = %s"
            val = ("YES", executed_contracts[int(i)].contract_id)
            self.mycursor.execute(sql, val)

        sql = "UPDATE users SET paid_points=%s, paid_cash=%s WHERE user_id = %s"
        val = (user.paid_points + paid_points_to_add, user.paid_cash + cash_to_add, user.id)
        self.mycursor.execute(sql, val)
        user.paid_points = user.paid_points + paid_points_to_add
        user.cash = user.paid_cash + cash_to_add
        self.mydb.commit()

        print("Executed contracts paid out successfully")



